from openpyxl import load_workbook
import re
""" Работа с exel файлами
С фриланса:
Нужно написать скрипт, который ищет пустые ячейки и заменяет их на определенные значения

"""

wb = load_workbook('exel_table.xlsx')
ws = wb['Лист1']

def search_empty_cells():
    """ функция ищет пустые значения и применяет к ним функцию insert_in_cells"""
    for i in ws:
        for j in i:
            cell = str(j)[-3:-1] # имя ячейки в str формате
            print('cell = ',cell, ws[cell].value)
            if ws[cell].value == None:
                if cell[0] == 'B':
                    insert_into_cell(cell, 'пустое имя')
                elif cell[0] == 'C':
                    insert_into_cell(cell, 'нет фамилии')
                elif cell[0] == 'D':
                    insert_into_cell(cell, 'нет телефона')
                else:
                    insert_into_cell(cell, 'нет почты')


def insert_into_cell(cell: str, data: str):
    """ вставляет в ячейку cell значение data"""
    ws[cell] = data
    for i in ws.values:
        print(i)
    wb.save('exel_table2.xlsx')

if __name__ == '__main__':
    search_empty_cells()




